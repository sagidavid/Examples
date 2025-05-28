# Load required libraries
import os
import pandas as pd
import glob
from openpyxl import load_workbook

# Specify path
username = os.path.expanduser("~").split('\\')[-1].rstrip('_adm')
root = 'C:/Users/'+username+'/Department of Health and Social Care/NW025 - Medicines/1_Continuity of Supply/30_Silver/09. Jan 25 GTM/03. Responses'

#%% Import Silver only data

input_path_silver_only = os.path.join(root, "Silver only/")
all_files_silver_only = glob.glob(input_path_silver_only + "/*.xlsx")
output_precursor = []

for filename in all_files_silver_only:
    wb_w = load_workbook(filename)
    wb_s = wb_w.active

    # Read MAH value
    MAH_name = wb_s.cell(row = 9, column = 3).value
    
    # Specify the starting row
    start_row = 15

    # Read the data
    data = []
    for row in wb_s.iter_rows(min_row=start_row, values_only=True):
        data.append(row)

    # Create dataframe
    df = pd.DataFrame(data)
    df = df.dropna(axis = 0, how = 'all')
    df['File.Path'] = filename
    df['File.Name'] = df['File.Path'].apply(os.path.basename)
    df['Date.Str'] = df['File.Name'].str.extract(r'_(\d{6})_')
    df['Return.Date'] = pd.to_datetime(df['Date.Str'], format='%y%m%d')

    # Insert MAH
    df.insert(0, "Marketing Authorisation Holder", MAH_name)

    output_precursor.append(df)

output_silver_only = pd.concat(output_precursor, axis=0, ignore_index=True, sort=False)
output_silver_only.to_excel(os.path.join(root, 'Collated Templates_Silver_Only.xlsx'), index=False)

#%% Import Silver and Winter data

input_path_silver_winter = os.path.join(root, "Silver and Winter")
all_files_silver_winter = glob.glob(input_path_silver_winter + "/*.xlsx")
output_precursor = []

for filename in all_files_silver_winter:
    wb_w = load_workbook(filename)
    wb_s = wb_w.active

    # Read MAH value
    MAH_name = wb_s.cell(row = 9, column = 3).value
    
    # Specify the starting row
    start_row = 15

    # Read the data
    data = []
    for row in wb_s.iter_rows(min_row=start_row, values_only=True):
        data.append(row)

    # Create dataframe
    df = pd.DataFrame(data)
    df = df.dropna(axis = 0, how = 'all')
    df['File.Path'] = filename
    df['File.Name'] = df['File.Path'].apply(os.path.basename)
    df['Date.Str'] = df['File.Name'].str.extract(r'_(\d{6})_')
    df['Return.Date'] = pd.to_datetime(df['Date.Str'], format='%y%m%d')
    
    # Insert MAH
    df.insert(0, "Marketing Authorisation Holder", MAH_name)

    output_precursor.append(df)

output_silver_winter = pd.concat(output_precursor, axis=0, ignore_index=True, sort=False)
output_silver_winter.to_excel(os.path.join(root, 'Collated Templates_Silver_Winter.xlsx'), index=False)

#%% Import Additional data

input_path_additional = os.path.join(root, "Additional")
all_files_additional = glob.glob(input_path_additional + "/*.xlsx")
output_precursor = []

for filename in all_files_additional:
    wb_w = load_workbook(filename)
    wb_s = wb_w.active

    # Read MAH value
    MAH_name = wb_s.cell(row = 9, column = 3).value
    
    # Specify the starting row
    start_row = 15

    # Read the data
    data = []
    for row in wb_s.iter_rows(min_row=start_row, values_only=True):
        data.append(row)

    # Create dataframe
    df = pd.DataFrame(data)
    df = df.dropna(axis = 0, how = 'all')
    df['File.Path'] = filename
    df['File.Name'] = df['File.Path'].apply(os.path.basename)
    df['Date.Str'] = df['File.Name'].str.extract(r'_(\d{6})_')
    df['Return.Date'] = pd.to_datetime(df['Date.Str'], format='%y%m%d')
    
    # Insert MAH
    df.insert(0, "Marketing Authorisation Holder", MAH_name)

    output_precursor.append(df)

output_additional = pd.concat(output_precursor, axis=0, ignore_index=True, sort=False)
output_additional.to_excel(os.path.join(root, 'Collated Templates_Additional.xlsx'), index=False)
