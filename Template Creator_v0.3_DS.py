# Load required libraries
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Protection

# Specify path
## Update the root depending on how you sync the folder
username = os.path.expanduser("~").split('\\')[-1].rstrip('_adm')
root = 'C:/Users/'+username+'/OneDrive - Department of Health and Social Care/1_Continuity of Supply/30_Silver/04. Silver Collation Folders/Python/01i. Templates to send 250528'

# Specify file names
input_data_file = '250528_Thistle_Jun 25_GTM_v1_Template.xlsx'
input_data_sheet = 'Consolidated List'

template_silver = 'Proposed GTM spreadsheet MST 250528.xlsx'
template_sheet = 'GTM Template'
output_name = 'GTM_MST'

# Specify protected password
password = "DHSC2024"

#%% Create Data
# Load data
input_data_raw = pd.read_excel(os.path.join(root, 'Data', input_data_file), sheet_name = input_data_sheet)

#%% Clean Data
columns = ['Authorisation.Number', 'Licensed.Product.Name', 'Drug.Substance', 'Strength', 
           'Presentation', 'Legal.Status.Type', 'Pack.Size', 'Silver.Actively.Marketed', 'Authorisation.Holder.Company.Name']

input_data_marketed = input_data_raw[columns]
input_data_marketed['Silver.Actively.Marketed'].replace('Not sent', 'N', inplace=True)
input_data_marketed = input_data_marketed[input_data_marketed['Silver.Actively.Marketed'] != "N"]
input_data_marketed = input_data_marketed.drop(columns = ["Silver.Actively.Marketed"])  
input_data_marketed = input_data_marketed.rename(columns = {'Authorisation.Holder.Company.Name' : 'MAH'})

# Group data by MAH
grouped_data_silver = input_data_marketed.groupby("MAH")

#%% Create templates (Silver Only)
# Process data for each MAH
for MAH, data in grouped_data_silver:

    # Load the template
    template_wb = load_workbook(os.path.join(root, 'Template', template_silver))
    
    # Select Sheet
    template_ws = template_wb[template_sheet]

    ## Insert date and MAH names
    template_ws.cell(row = 8, column = 3, value = datetime.today().strftime("%d-%m-%Y")).protection = Protection(locked=True)
    template_ws.cell(row = 9, column = 3, value = MAH).protection = Protection(locked=True)

    ## Remove MAH column
    data = data.drop(columns = ["MAH"])  

    ## Determine where to start writing data in the template
    start_row = 15
    start_column = 1

    ## Write each row of data into the template
    for i, row in enumerate(dataframe_to_rows(data, index = False, header = False), start=start_row):
        for j, value in enumerate(row, start = start_column):
            template_ws.cell(row = i, column = j, value = value).protection = Protection(locked=True)

    ## Protection Settings
    template_ws.protection.set_password(password)
    prot = template_ws.protection
    prot.formatColumns = False

    # Save the file for this teacher
    template_wb.save(os.path.join(root,"output", f'{output_name}-{MAH}.xlsx'))
    print(f"File saved for MAH: {MAH}")
